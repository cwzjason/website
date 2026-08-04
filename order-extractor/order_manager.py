# 订单信息提取与汇总系统 - 订单数据管理模块

import os
import json
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional

from config import ORDER_FIELDS_A, ORDER_FIELDS_B, DATA_FILE

# 合并两套字段体系，确保Excel能存储所有平台的订单
ALL_COLUMNS = ORDER_FIELDS_A + [f for f in ORDER_FIELDS_B if f not in ORDER_FIELDS_A]
# 平台来源作为元数据字段，需要单独加入（用于数据筛选）
ALL_COLUMNS = ALL_COLUMNS + ["平台来源"]


class OrderManager:
    """订单数据管理器，负责订单的增删改查和Excel导出"""
    
    def __init__(self, data_file: str = DATA_FILE):
        self.data_file = data_file
        self.columns = ALL_COLUMNS + ["_source_file", "_processed_at", "_id", "_platform_type"]
        self._ensure_data_file()
    
    def _ensure_data_file(self):
        """确保数据文件存在"""
        if not os.path.exists(self.data_file):
            df = pd.DataFrame(columns=self.columns)
            df.to_excel(self.data_file, index=False, engine='openpyxl')
    
    def _load_data(self) -> pd.DataFrame:
        """加载数据"""
        try:
            df = pd.read_excel(self.data_file, engine='openpyxl')
            # 确保列完整
            for col in self.columns:
                if col not in df.columns:
                    df[col] = ""
                else:
                    df[col] = df[col].fillna("")
            return df
        except Exception:
            return pd.DataFrame(columns=self.columns)
    
    # 必须作为文本存储的列（避免长数字被转成浮点数/科学计数法）
    _TEXT_COLS = {
        '采购单编号', '需求单编号', '采购单号', '合约单号', '供应商编码',
        '收货人联系方式', '联系电话', '序号', '货号',
    }

    def _save_data(self, df: pd.DataFrame):
        """保存数据到Excel"""
        # 强制将数字类字段转为文本存储，避免浮点数问题
        for col in self._TEXT_COLS:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() != '' and str(x).replace('.','',1).replace('-','',1).isdigit() else (str(x).strip() if pd.notna(x) and str(x).strip() != '' else ''))

        with pd.ExcelWriter(self.data_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='订单汇总')

            worksheet = writer.sheets['订单汇总']

            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment

            for i, col_name in enumerate(self.columns):
                col_letter = get_column_letter(i + 1)
                if len(col_name) > 10:
                    worksheet.column_dimensions[col_letter].width = max(len(col_name) * 2, 20)
                elif len(col_name) > 6:
                    worksheet.column_dimensions[col_letter].width = 14
                else:
                    worksheet.column_dimensions[col_letter].width = 10

            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for cell in worksheet[1]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def add_order(self, order_data: Dict) -> Dict:
        """添加订单"""
        df = self._load_data()
        
        # 生成唯一ID
        order_id = f"ORD{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
        order_data["_id"] = order_id
        
        # 确保所有字段存在
        for col in self.columns:
            if col not in order_data:
                order_data[col] = ""
        
        # 按列顺序创建新行
        new_row = {col: order_data.get(col, "") for col in self.columns}
        new_df = pd.DataFrame([new_row])
        df = pd.concat([df, new_df], ignore_index=True)
        
        self._save_data(df)
        return new_row
    
    def get_all_orders(self, page: int = 1, page_size: int = 50, 
                       search: str = "", platform: str = "") -> Dict:
        """获取订单列表（支持分页和筛选）"""
        df = self._load_data()
        
        # 筛选
        if platform:
            if platform == "国铁商城":
                # 国铁商城订单 = 得力 + 阳采 + 德致 + 其他
                df = df[df["平台来源"].isin(["得力", "阳采", "德致", "其他"])]
            else:
                df = df[df["平台来源"] == platform]
        
        if search:
            # 全文搜索
            mask = df.apply(
                lambda row: row.astype(str).str.contains(search, case=False, na=False).any(), 
                axis=1
            )
            df = df[mask]
        
        total = len(df)
        
        # 分页
        start = (page - 1) * page_size
        end = start + page_size
        page_df = df.iloc[start:end]
        
        # 替换NaN
        page_df = page_df.where(page_df.notna(), "")
        
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "data": page_df.to_dict(orient="records")
        }
    
    def get_order_by_id(self, order_id: str) -> Optional[Dict]:
        """根据ID获取订单"""
        df = self._load_data()
        matches = df[df["_id"] == order_id]
        if len(matches) > 0:
            return matches.iloc[0].where(matches.iloc[0].notna(), "").to_dict()
        return None
    
    def update_order(self, order_id: str, updates: Dict) -> Optional[Dict]:
        """更新订单信息"""
        df = self._load_data()
        idx = df[df["_id"] == order_id].index
        
        if len(idx) == 0:
            return None
        
        for key, value in updates.items():
            if key in self.columns and key != "_id":
                df.loc[idx[0], key] = value
        
        self._save_data(df)
        return self.get_order_by_id(order_id)
    
    def delete_order(self, order_id: str) -> bool:
        """删除订单"""
        df = self._load_data()
        before = len(df)
        df = df[df["_id"] != order_id]
        
        if len(df) < before:
            self._save_data(df)
            return True
        return False
    
    def get_statistics(self) -> Dict:
        """获取统计数据"""
        df = self._load_data()
        
        stats = {
            "订单总数": len(df),
            "平台分布": {},
            "总金额": 0.0,
            "国铁商城金额": 0.0,
            "齐心金额": 0.0,
            "日期范围": {"开始": "", "结束": ""},
            "最近订单": []
        }
        
        if len(df) == 0:
            return stats
        
        # 平台分布
        platform_counts = df["平台来源"].value_counts().to_dict()
        stats["平台分布"] = {k: int(v) for k, v in platform_counts.items() if k}
        
        # 安全转浮点数
        def safe_float(x):
            try:
                if x and str(x).replace('.','').replace('-','').isdigit():
                    return float(x)
            except:
                pass
            return 0.0
        
        # 国铁商城订单金额：平台为得力/阳采/德致/其他，取"合计（含税）"列
        guotie_mask = df["平台来源"].isin(["得力", "阳采", "德致", "其他"])
        if "合计（含税）" in df.columns:
            guotie_amount = df.loc[guotie_mask, "合计（含税）"].apply(safe_float).sum()
        else:
            guotie_amount = 0.0
        
        # 齐心订单金额：平台为齐心，取"金额合计（小写）"列
        qixin_mask = df["平台来源"] == "齐心"
        if "金额合计（小写）" in df.columns:
            qixin_amount = df.loc[qixin_mask, "金额合计（小写）"].apply(safe_float).sum()
        else:
            qixin_amount = 0.0
        
        stats["国铁商城金额"] = round(guotie_amount, 2)
        stats["齐心金额"] = round(qixin_amount, 2)
        stats["总金额"] = round(guotie_amount + qixin_amount, 2)
        
        # 日期范围
        try:
            dates = df["下单日期"].dropna()
            dates = dates[dates != ""]
            if len(dates) > 0:
                sorted_dates = sorted(dates)
                stats["日期范围"]["开始"] = str(sorted_dates[0])
                stats["日期范围"]["结束"] = str(sorted_dates[-1])
        except:
            pass
        
        # 最近订单
        recent = df.sort_values("_processed_at", ascending=False).head(5)
        stats["最近订单"] = recent.where(recent.notna(), "").to_dict(orient="records")
        
        return stats
    
    def export_excel(self, output_path: str = None, platform: str = "all") -> str:
        """导出Excel文件 - 按平台分Sheet，各自独立列模板，不合并
        
        platform 参数：
          - "all": 导出全部（国铁商城 + 齐心各一个Sheet）
          - "guotie": 仅导出 国铁商城订单
          - "qixin": 仅导出 齐心订单
        """
        import tempfile

        if output_path is None:
            prefix = "订单汇总"
            if platform == "guotie":
                prefix = "国铁商城订单"
            elif platform == "qixin":
                prefix = "齐心订单"
            output_path = os.path.join(tempfile.gettempdir(),
                                        f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        df = self._load_data()

        # 国铁商城（得力/阳采/德致/其他）显示列
        GUOTIE_DISPLAY_COLS = [
            '采购单编号', '需求单编号', '审批通过时间',
            '收料单位', '所属路局',
            '供应商名称', '供应商编码', '下单人',
            '商品名称', '品牌', '数量',
            '单价（含税）', '合计（含税）', '合计金额',
            '收货人', '收货人联系方式', '收货地址',
            '发票备注', '订单备注',
        ]

        # 齐心显示列
        QIXIN_DISPLAY_COLS = [
            '采购单号', '发货日期',
            '客户名称', '客户地址',
            '收货人', '联系人', '联系电话',
            '送方名称', '收货地址',
            '合约单号', '经办人',
            '序号', '货号', '产品名称',
            '单位', '数量', '单价', '金额',
            '金额合计（小写）', '备注',
        ]

        # 金额列
        GUOTIE_AMOUNT_COLS = {'合计（含税）', '合计金额', '单价（含税）'}
        QIXIN_AMOUNT_COLS = {'单价', '金额', '金额合计（小写）'}

        # 需要固定为文本格式的列（订单号等长数字）
        GUOTIE_TEXT_COLS = {'采购单编号', '需求单编号', '供应商编码', '收货人联系方式'}
        QIXIN_TEXT_COLS = {'采购单号', '合约单号', '联系电话', '序号', '货号'}

        def _ensure_str(val):
            """确保值为纯文本字符串，不保留浮点痕迹"""
            if pd.isna(val) or str(val).strip() == '':
                return ''
            s = str(val).strip()
            try:
                return str(int(float(s)))
            except (ValueError, TypeError):
                return s

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

            # --- 国铁商城订单 ---
            if platform in ("all", "guotie"):
                guotie_df = df[df['平台来源'].isin(['得力', '阳采', '德致', '其他'])].copy()
                guotie_cols = [c for c in GUOTIE_DISPLAY_COLS if c in guotie_df.columns]
                if len(guotie_df) > 0 and guotie_cols:
                    guotie_export = guotie_df[guotie_cols].copy()
                    for col in guotie_export.columns:
                        if col in GUOTIE_TEXT_COLS:
                            guotie_export[col] = guotie_export[col].apply(_ensure_str)

                    guotie_export.to_excel(writer, index=False, sheet_name='国铁商城订单')
                    self._format_sheet(
                        writer.sheets['国铁商城订单'],
                        guotie_cols,
                        GUOTIE_AMOUNT_COLS,
                        GUOTIE_TEXT_COLS,
                    )

            # --- 齐心订单 ---
            if platform in ("all", "qixin"):
                qixin_df = df[df['平台来源'] == '齐心'].copy()
                qixin_cols = [c for c in QIXIN_DISPLAY_COLS if c in qixin_df.columns]
                if len(qixin_df) > 0 and qixin_cols:
                    qixin_export = qixin_df[qixin_cols].copy()
                    for col in qixin_export.columns:
                        if col in QIXIN_TEXT_COLS:
                            qixin_export[col] = qixin_export[col].apply(_ensure_str)

                    qixin_export.to_excel(writer, index=False, sheet_name='齐心订单')
                    self._format_sheet(
                        writer.sheets['齐心订单'],
                        qixin_cols,
                        QIXIN_AMOUNT_COLS,
                        QIXIN_TEXT_COLS,
                    )

        return output_path
    
    @staticmethod
    def _format_sheet(worksheet, columns, amount_cols, number_cols):
        """格式化工作表样式"""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        
        # 表头样式
        header_font = Font(bold=True, size=11, color='FFFFFF', name='微软雅黑')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 数据行样式
        data_font = Font(size=10, name='微软雅黑')
        data_alignment = Alignment(vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')
        amount_alignment = Alignment(horizontal='right', vertical='center')
        
        # 边框
        thin_border = Border(
            left=Side(style='thin', color='D0D5DD'),
            right=Side(style='thin', color='D0D5DD'),
            top=Side(style='thin', color='D0D5DD'),
            bottom=Side(style='thin', color='D0D5DD'),
        )
        
        # 斑马条纹颜色
        stripe_fill = PatternFill(start_color='F8F9FF', end_color='F8F9FF', fill_type='solid')
        
        # 设置列宽
        for i, col_name in enumerate(columns):
            col_letter = get_column_letter(i + 1)
            col_name_len = len(col_name)
            if col_name_len > 12:
                width = min(col_name_len * 2.2, 40)
            elif col_name_len > 6:
                width = 16
            else:
                width = 12
            worksheet.column_dimensions[col_letter].width = width
        
        # 设置行高
        worksheet.row_dimensions[1].height = 28
        
        # 格式化表头
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 格式化数据行
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                                           min_col=1, max_col=len(columns)), start=2):
            for col_idx, cell in enumerate(row):
                col_name = columns[col_idx]
                
                cell.font = data_font
                cell.border = thin_border
                
                # 金额列右对齐
                if col_name in amount_cols:
                    cell.alignment = amount_alignment
                    if cell.value and str(cell.value).strip():
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '#,##0.00'
                        except (ValueError, TypeError):
                            pass
                # 数字列居中，文本格式避免科学计数法
                elif col_name in number_cols:
                    cell.alignment = center_alignment
                    cell.number_format = '@'  # 文本格式
                else:
                    cell.alignment = data_alignment
                
                # 斑马条纹
                if row_idx % 2 == 0:
                    cell.fill = stripe_fill
        
        # 冻结首行
        worksheet.freeze_panes = 'A2'
        
        # 自动筛选
        worksheet.auto_filter.ref = f'A1:{get_column_letter(len(columns))}{worksheet.max_row}'


# 全局单例
order_manager = OrderManager()
