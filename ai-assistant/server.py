"""
埠勤商贸 AI 智能助手 - 后端服务
DeepSeek 多模态 + SQLite 数据存储
"""
import os
import json
import base64
import io
import sqlite3
import re
from datetime import datetime
from pathlib import Path

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook
from PIL import Image
import requests

app = Flask(__name__)
CORS(app)

# ── 配置 ──────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "data.db"
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
DEEPSEEK_BASE_URL = "https://api.deepseek.com"
DEEPSEEK_MODEL = "deepseek-chat"

# ── 数据库初始化 ───────────────────────────────────
def init_db():
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS daily_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            category TEXT DEFAULT '',
            item TEXT DEFAULT '',
            amount REAL DEFAULT 0,
            remark TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS chat_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            role TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# ── 工具函数 ───────────────────────────────────────
def parse_excel(file_bytes: bytes, filename: str) -> str:
    """解析 Excel 文件，返回 Markdown 表格文本"""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"## 工作表: {sheet_name}")
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                all_text.append("(空表)")
                continue
            for i, row in enumerate(rows):
                cells = [str(c) if c is not None else "" for c in row]
                all_text.append(" | ".join(cells))
                if i == 0 and len(rows) > 1:
                    all_text.append("|".join(["---"] * len(cells)))
        return "\n".join(all_text)
    except Exception as e:
        return f"[Excel 解析失败: {str(e)}]"

def parse_image(file_bytes: bytes) -> dict:
    """解析图片基本信息"""
    try:
        img = Image.open(io.BytesIO(file_bytes))
        info = {
            "格式": img.format,
            "尺寸": f"{img.width}x{img.height}",
            "模式": img.mode,
        }
        # 尝试读取 EXIF
        exif_data = img.getexif()
        if exif_data:
            for tag_id, value in exif_data.items():
                from PIL.ExifTags import TAGS
                tag_name = TAGS.get(tag_id, tag_id)
                if tag_name in ("DateTime", "Make", "Model"):
                    info[tag_name] = str(value)
        text_lines = [f"{k}: {v}" for k, v in info.items()]
        return {"text": "\n".join(text_lines), "info": info}
    except Exception as e:
        return {"text": f"[图片解析失败: {str(e)}]", "info": {}}

def call_deepseek(messages: list, tools: list = None) -> dict:
    """调用 DeepSeek API"""
    if not DEEPSEEK_API_KEY:
        return {"error": "未配置 DEEPSEEK_API_KEY 环境变量"}

    headers = {
        "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": DEEPSEEK_MODEL,
        "messages": messages,
        "temperature": 0.3,
        "max_tokens": 2000,
    }
    if tools:
        payload["tools"] = tools
        payload["tool_choice"] = "auto"

    resp = requests.post(
        f"{DEEPSEEK_BASE_URL}/v1/chat/completions",
        headers=headers,
        json=payload,
        timeout=60
    )
    if resp.status_code != 200:
        return {"error": f"API 错误 {resp.status_code}: {resp.text[:300]}"}
    return resp.json()

# ── 数据库操作函数 ─────────────────────────────────
def insert_record(date_str: str, category: str, item: str, amount: float, remark: str) -> dict:
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()
    c.execute(
        "INSERT INTO daily_records (date, category, item, amount, remark) VALUES (?,?,?,?,?)",
        (date_str, category, item, amount, remark)
    )
    record_id = c.lastrowid
    conn.commit()
    conn.close()
    return {"id": record_id, "date": date_str, "category": category, "item": item, "amount": amount, "remark": remark}

def query_records(date_str: str = None, limit: int = 20) -> list:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    if date_str:
        c.execute("SELECT * FROM daily_records WHERE date=? ORDER BY id DESC LIMIT ?", (date_str, limit))
    else:
        c.execute("SELECT * FROM daily_records ORDER BY id DESC LIMIT ?", (limit,))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows

def get_stats() -> dict:
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as total, SUM(amount) as total_amount FROM daily_records")
    row = c.fetchone()
    c.execute("SELECT category, COUNT(*) as cnt, SUM(amount) as total FROM daily_records GROUP BY category ORDER BY total DESC LIMIT 10")
    categories = [{"category": r[0], "count": r[1], "total": r[2]} for r in c.fetchall()]
    conn.close()
    return {
        "total_records": row[0] or 0,
        "total_amount": row[1] or 0,
        "categories": categories
    }

# ── System Prompt ─────────────────────────────────
SYSTEM_PROMPT = """你是埠勤商贸的 AI 数据助手「小埠」。你的核心职责是帮助用户管理日常流水记录。

## 你的能力
1. **新增流水记录**: 用户说"帮我记一笔"、"新增一条流水"等，你需要提取：日期、类别、项目、金额、备注
2. **查询流水**: 用户问"今天记了哪些"、"这个月花了多少"等，调工具查询
3. **分析统计**: 用户问分类汇总、总计等，调工具获取统计

## 数据表字段
- date: 日期，格式 YYYY-MM-DD（未指定则用今天）
- category: 类别（如：餐饮、交通、采购、办公、其他）
- item: 具体项目名称
- amount: 金额（数字）
- remark: 备注说明

## 规则
- 日期未指定时默认今天（{today}）
- 金额必须是正数
- 回复简洁友好，用中文
- 插入成功后返回"已记录：[日期] [类别] [项目] ¥[金额]"
- 如果用户上传了 Excel/图片，先仔细阅读内容再回复

当前日期：{today}
"""

# ── Function Definitions ───────────────────────────
TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "add_record",
            "description": "新增一条每日流水记录",
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "日期，格式 YYYY-MM-DD"},
                    "category": {"type": "string", "description": "类别，如：餐饮、交通、采购、办公、其他"},
                    "item": {"type": "string", "description": "具体项目名称"},
                    "amount": {"type": "number", "description": "金额"},
                    "remark": {"type": "string", "description": "备注说明"}
                },
                "required": ["date", "category", "item", "amount"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "search_records",
            "description": "查询流水记录",
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "查询日期，格式 YYYY-MM-DD，不传则查全部"},
                    "limit": {"type": "integer", "description": "返回条数，默认20"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_statistics",
            "description": "获取流水统计信息，包括总数、总金额、分类汇总",
            "parameters": {"type": "object", "properties": {}}
        }
    }
]

# ── API 路由 ────────────────────────────────────────
@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "model": DEEPSEEK_MODEL, "has_api_key": bool(DEEPSEEK_API_KEY)})

@app.route("/api/chat", methods=["POST"])
def chat():
    """主对话接口，支持纯文本、图片、Excel 文件"""
    try:
        text = request.form.get("message", "")
        file = request.files.get("file")

        if not text and not file:
            return jsonify({"error": "请提供消息或文件"}), 400

        today = datetime.now().strftime("%Y-%m-%d")
        today_time = datetime.now().strftime("%Y-%m-%d %H:%M")

        # 构建消息
        messages = []
        user_content = []

        # 处理文件
        file_info = ""
        if file:
            file_bytes = file.read()
            filename = file.filename or ""
            ext = Path(filename).suffix.lower() if filename else ""

            if ext in (".xlsx", ".xls"):
                excel_text = parse_excel(file_bytes, filename)
                file_info = f"\n\n【用户上传了 Excel 文件: {filename}】\n{excel_text}"
            elif ext in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"):
                img_result = parse_image(file_bytes)
                file_info = f"\n\n【用户上传了图片: {filename}】\n{img_result['text']}"
                # 图片 base64 给 DeepSeek vision
                b64 = base64.b64encode(file_bytes).decode("utf-8")
                mime = f"image/{ext.lstrip('.')}"
                if ext == ".jpg":
                    mime = "image/jpeg"
                user_content.append({
                    "type": "image_url",
                    "image_url": {"url": f"data:{mime};base64,{b64}"}
                })
            else:
                try:
                    text_content = file_bytes.decode("utf-8")
                    file_info = f"\n\n【用户上传了文件: {filename}】\n{text_content}"
                except:
                    file_info = f"\n\n【用户上传了文件: {filename}（二进制文件，无法解析文本内容）】"

        # 系统提示
        messages.append({
            "role": "system",
            "content": SYSTEM_PROMPT.format(today=today)
        })

        # 加载历史对话（正确顺序：旧 → 新）
        history = load_chat_history(10)
        for h in history:
            messages.append({"role": h["role"], "content": h["content"]})

        # 构建当前用户消息
        if user_content:
            user_content.insert(0, {"type": "text", "text": text + file_info if text else file_info})
            messages.append({"role": "user", "content": user_content})
        else:
            full_text = text + file_info if file_info else text
            messages.append({"role": "user", "content": full_text})

        # 保存用户消息（用于历史记录）
        save_chat("user", text or f"[上传文件: {file.filename if file else '未知'}]")

        # 第一次调用 DeepSeek（可能触发 function calling）
        result = call_deepseek(messages, tools=TOOLS)

        if "error" in result:
            return jsonify({"reply": f"AI 服务暂时不可用: {result['error']}", "actions": []}), 500

        choice = result["choices"][0]
        msg = choice["message"]

        actions = []

        # 处理 function calling
        if msg.get("tool_calls"):
            for tc in msg["tool_calls"]:
                func = tc["function"]
                args = json.loads(func["arguments"])
                tool_result = ""

                if func["name"] == "add_record":
                    rec = insert_record(
                        args.get("date", today),
                        args.get("category", "其他"),
                        args.get("item", ""),
                        args.get("amount", 0),
                        args.get("remark", "")
                    )
                    tool_result = json.dumps({"success": True, "record": rec}, ensure_ascii=False)
                    actions.append({"type": "add", "record": rec})

                elif func["name"] == "search_records":
                    records = query_records(args.get("date"), args.get("limit", 20))
                    tool_result = json.dumps(records, ensure_ascii=False, default=str)

                elif func["name"] == "get_statistics":
                    stats = get_stats()
                    tool_result = json.dumps(stats, ensure_ascii=False)

            # 将工具调用结果发回 DeepSeek 生成最终回复
            messages.append(msg)
            messages.append({
                "role": "tool",
                "tool_call_id": tc.get("id", "call_1"),
                "content": tool_result
            })

            result2 = call_deepseek(messages)
            if "error" not in result2:
                reply = result2["choices"][0]["message"]["content"]
            else:
                reply = f"操作完成！{tool_result[:200]}"
        else:
            reply = msg.get("content", "")

        # 保存 AI 回复
        save_chat("assistant", reply)

        return jsonify({
            "reply": reply,
            "actions": actions,
            "time": today_time
        })

    except Exception as e:
        return jsonify({"reply": f"处理出错: {str(e)}", "actions": []}), 500

@app.route("/api/records", methods=["GET"])
def get_records():
    date = request.args.get("date")
    limit = int(request.args.get("limit", 50))
    records = query_records(date, limit)
    return jsonify(records)

@app.route("/api/stats", methods=["GET"])
def stats():
    return jsonify(get_stats())

# ── 聊天历史 ────────────────────────────────────────
def save_chat(role: str, content: str):
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()
    c.execute("INSERT INTO chat_history (role, content) VALUES (?,?)", (role, content))
    conn.commit()
    conn.close()

def load_chat_history(limit: int = 10) -> list:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT role, content FROM chat_history ORDER BY id DESC LIMIT ?", (limit,))
    rows = [{"role": r["role"], "content": r["content"]} for r in c.fetchall()]
    conn.close()
    return list(reversed(rows))

# ── 静态文件（前端 AI 聊天页面）─────────────────────
@app.route("/")
def serve_frontend():
    return send_from_directory(str(BASE_DIR), "chat.html")

# ── 启动 ────────────────────────────────────────────
if __name__ == "__main__":
    print(f"🚀 埠勤 AI 助手启动")
    print(f"   数据库: {DB_PATH}")
    print(f"   上传目录: {UPLOAD_DIR}")
    print(f"   DeepSeek API Key: {'已配置' if DEEPSEEK_API_KEY else '❌ 未配置'}")
    app.run(host="0.0.0.0", port=8085, debug=False)
